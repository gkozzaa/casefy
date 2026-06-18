import json
import re
import os
from openpyxl import load_workbook

UPLOADS_DIR = "/mnt/user-data/uploads"
METADATA_FILE = "/mnt/user-data/outputs/pm_interviews_metadata.xlsx"
OUTPUT_JSONL = "/mnt/user-data/outputs/pm_rag_chunks.jsonl"

# Quantas falas do candidato por chunk
TURNS_PER_CHUNK = 3
# Overlap: reutiliza N turnos do chunk anterior para contexto
OVERLAP_TURNS = 1

SPEAKER_PATTERNS = [
    r"^(Interviewer|Interviewee|Candidate|CAndidate|SPK_\d+|Host|Guest)\s*$"
]

def load_metadata():
    wb = load_workbook(METADATA_FILE, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows

def parse_transcript(filepath):
    """
    Retorna lista de dicts: [{speaker, text}, ...]
    Agrupa linhas consecutivas do mesmo speaker num único turno.
    """
    with open(filepath, encoding="utf-8") as f:
        lines = [l.rstrip() for l in f.readlines()]

    turns = []
    current_speaker = None
    current_lines = []

    for line in lines:
        is_speaker = any(re.match(p, line.strip()) for p in SPEAKER_PATTERNS)
        if is_speaker:
            if current_speaker and current_lines:
                text = " ".join(current_lines).strip()
                if text:
                    turns.append({"speaker": current_speaker, "text": text})
            current_speaker = line.strip()
            current_lines = []
        elif line.strip():
            current_lines.append(line.strip())

    # flush último turno
    if current_speaker and current_lines:
        text = " ".join(current_lines).strip()
        if text:
            turns.append({"speaker": current_speaker, "text": text})

    return turns

def classify_speaker(speaker_label, meta):
    """Retorna 'candidate' ou 'interviewer'"""
    interviewer_labels = {"Interviewer", "SPK_1", "Host"}
    candidate_labels   = {"Candidate", "CAndidate", "SPK_2", "Guest", "Interviewee"}
    if speaker_label in interviewer_labels:
        return "interviewer"
    if speaker_label in candidate_labels:
        return "candidate"
    return "unknown"

def build_chunks(turns, meta):
    """
    Estratégia: chunk por N turnos consecutivos do candidato,
    mantendo contexto da pergunta do entrevistador antes.
    """
    chunks = []
    chunk_id = 0

    # índices de turnos do candidato
    candidate_indices = [
        i for i, t in enumerate(turns)
        if classify_speaker(t["speaker"], meta) == "candidate"
    ]

    # janela deslizante com overlap
    step = max(1, TURNS_PER_CHUNK - OVERLAP_TURNS)
    windows = []
    i = 0
    while i < len(candidate_indices):
        windows.append(candidate_indices[i:i + TURNS_PER_CHUNK])
        i += step

    for window in windows:
        if not window:
            continue

        first_cand_idx = window[0]
        last_cand_idx  = window[-1]

        # pegar pergunta do entrevistador imediatamente antes do primeiro turno
        question_text = ""
        if first_cand_idx > 0:
            prev = turns[first_cand_idx - 1]
            if classify_speaker(prev["speaker"], meta) == "interviewer":
                question_text = prev["text"]

        # concatenar respostas do candidato na janela
        candidate_texts = [turns[idx]["text"] for idx in window if idx < len(turns)]
        response_text = " ".join(candidate_texts)

        # contexto: pergunta + resposta
        full_text = f"[Interviewer]: {question_text}\n[Candidate]: {response_text}" if question_text else f"[Candidate]: {response_text}"

        chunk = {
            "chunk_id": f"{meta['video_id']}_{chunk_id:03d}",
            "video_id": meta["video_id"],
            "filename": meta["filename"],
            "titulo": meta["titulo"],
            "empresa_alvo": meta["empresa_alvo"],
            "tipo_entrevista": meta["tipo_entrevista"],
            "produto_caso": meta["produto_caso"],
            "dimensoes_cobertas": [d.strip() for d in str(meta["dimensoes_cobertas"]).split(";")],
            "candidato_nivel": meta["candidato_nivel"],
            "qualidade_resposta": meta["qualidade_resposta"],
            "fonte_canal": meta["fonte_canal"],
            "notas_rag": meta["notas_rag"],
            "turn_range": [first_cand_idx, last_cand_idx],
            "text": full_text,
            "word_count": len(full_text.split()),
        }
        chunks.append(chunk)
        chunk_id += 1

    return chunks

def main():
    metadata = load_metadata()
    all_chunks = []

    for meta in metadata:
        filepath = os.path.join(UPLOADS_DIR, meta["filename"])
        if not os.path.exists(filepath):
            print(f"  SKIP (não encontrado): {meta['filename']}")
            continue

        turns = parse_transcript(filepath)
        chunks = build_chunks(turns, meta)
        all_chunks.extend(chunks)
        print(f"  {meta['video_id']} | {len(turns)} turnos → {len(chunks)} chunks | {meta['titulo'][:60]}")

    with open(OUTPUT_JSONL, "w", encoding="utf-8") as f:
        for chunk in all_chunks:
            f.write(json.dumps(chunk, ensure_ascii=False) + "\n")

    print(f"\nTotal: {len(all_chunks)} chunks → {OUTPUT_JSONL}")

    # preview do primeiro e último chunk
    if all_chunks:
        print("\n--- PRIMEIRO CHUNK ---")
        c = all_chunks[0]
        print(f"ID: {c['chunk_id']} | Dimensões: {c['dimensoes_cobertas']}")
        print(f"Texto ({c['word_count']} palavras):\n{c['text'][:400]}...")

        print("\n--- ÚLTIMO CHUNK ---")
        c = all_chunks[-1]
        print(f"ID: {c['chunk_id']} | Dimensões: {c['dimensoes_cobertas']}")
        print(f"Texto ({c['word_count']} palavras):\n{c['text'][:400]}...")

if __name__ == "__main__":
    main()
